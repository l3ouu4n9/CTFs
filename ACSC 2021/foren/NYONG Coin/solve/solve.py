#!/usr/bin/env python3

import openpyxl



xlsx_file1 = './20200715132932_20200816181652.xlsx'
wb_obj1 = openpyxl.load_workbook(xlsx_file1)
sheet1 = wb_obj1.active


data1 = []
for i, row in enumerate(sheet1.iter_rows(values_only=True)):
    if i == 0:
        pass
    else:
        a = str(row[0]) + row[1] + row[2] + row[3] + row[4] + row[5] + row[6]
        data1.append(a)

print("[+] Finish sheet 1")


xlsx_file2 = './f0015088.xlsx'
wb_obj2 = openpyxl.load_workbook(xlsx_file2)
sheet2 = wb_obj2.active


data2 = []
for i, row in enumerate(sheet2.iter_rows(values_only=True)):
    if i == 0:
        pass
    else:
        a = str(row[0]) + row[1] + row[2] + row[3] + row[4] + row[5] + row[6]
        data2.append(a)
        

print("[+] Finish sheet 2")

for i in range(0, len(data1)):
    if data1[i] != data2[i]:
        print(i)
        print("Data 1 is: ", data1[i])
        print("Data 2 is: ", data2[i])

"""
[+] Finish sheet 1
[+] Finish sheet 2
40990
Data 1 is:  2020-07-28 16:55:588d77a554-dc64-478c-b093-da4493a8534dNYONG***** USD6151.7 NYONGErlciooedxtiyotrtnzbsbdtbezsstrilfqbflbgoupvxpfzaicrwupuzfqilsrph
Data 2 is:  2020-07-28 16:55:588d77a554-dc64-478c-b093-da4493a8534cNYONG***** USD7151.7 NYONGErlciooedxtiyotrtnzbsbdtbezsstrilfqbflbgoupvxpfzaicrwupuzfqilsrph
flag is ACSC{8d77a554-dc64-478c-b093-da4493a8534d}
"""